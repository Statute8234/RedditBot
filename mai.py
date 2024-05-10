import openpyxl
import praw
import string
import random
import requests
import re
import urllib.request
from autocorrect import Speller
import concurrent.futures
import numpy as np
from PIL import Image
from io import BytesIO
import secrets
import time
# nltk
import nltk
from nltk import word_tokenize, pos_tag
from nltk.corpus import wordnet, stopwords
from nltk.tokenize import sent_tokenize
from nltk.stem import WordNetLemmatizer, PorterStemmer
from nltk.tokenize import RegexpTokenizer
import string
# tensorflow
from tensorflow.keras.preprocessing.text import Tokenizer
from tensorflow.keras.preprocessing.sequence import pad_sequences
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Embedding, LSTM, Dense
from tensorflow.keras.optimizers import Adam
from tensorflow.keras.applications.resnet50 import ResNet50, preprocess_input, decode_predictions
from tensorflow.keras.preprocessing import image as keras_image
from tensorflow.keras.preprocessing import image
# sklear
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder
# pandas
import pandas as pd
# Download NLTK resources
nltk.download('stopwords')
nltk.download('punkt')
nltk.download('averaged_perceptron_tagger')
nltk.download('wordnet')
# load model
resnet_model = ResNet50(weights='imagenet')
stopwords_list = set(stopwords.words('english'))
# open file
path = r"Data.xlsx"
expected_headers = ["Title", "Score", "Comments", "Awards", "Time", "Url", "Flair"]
try:
    workbook = openpyxl.load_workbook(path)
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    new_excel_file_path = 'Data.xlsx'
    [sheet.cell(row=1, column=i, value=header) for i, header in enumerate(expected_headers, start=1)]
    workbook.save(new_excel_file_path)
    print("Made a new file")
sheet = workbook.active
# Reddit Info
reddit = praw.Reddit(
    client_id="***",
    client_secret="***",
    user_agent="***",
)
current_row = 2
subreddit_name = "***"
subreddit = reddit.subreddit(subreddit_name)
num_posts = subreddit.subscribers
print(f"Looking through {num_posts} posts....")
# process submissions
def process_submission(submission):
    submission_data = (submission.title, submission.score, submission.num_comments, len(submission.all_awardings),submission.created_utc, submission.url, submission.link_flair_text)
    return submission_data
def update_sheet(current_row, submission_data):
    for col, value in enumerate(submission_data, start=1):
        sheet.cell(row=current_row, column=col, value=value)
NUM_THREADS = 4 # Number of threads (adjust based on your system)
with concurrent.futures.ThreadPoolExecutor(max_workers=NUM_THREADS) as executor:
    submission_data_list = list(executor.map(process_submission, reddit.subreddit(subreddit_name).new(limit=None)))
for current_row, submission_data in enumerate(submission_data_list, start=2):
    sheet_data = tuple(sheet.cell(row=current_row, column=col).value for col in range(1, len(expected_headers) + 1))
    if submission_data != sheet_data:
        update_sheet(current_row, submission_data)
workbook.save(path)
workbook.close()
# clean and correct text
worksheet = workbook["Sheet1"] 
ps = PorterStemmer()
def preprocess_text(text):
    words = word_tokenize(text)
    filtered_words = [ps.stem(w.lower()) for w in words if w.isalpha() and w.lower() not in stopwords_list and w not in string.punctuation and not w.isdigit()]
    return " ".join(filtered_words)
# Load your dataset
titles = [worksheet['A' + str(i)].value for i in range(2, worksheet.max_row + 1)]
preprocessed_titles = [preprocess_text(title) for title in titles]
tokenizer = Tokenizer()
tokenizer.fit_on_texts(preprocessed_titles)
vocab_size = len(tokenizer.word_index) + 1
sequences = tokenizer.texts_to_sequences(preprocessed_titles)
# Prepare input and output data for the model
input_sequences = []
output_sequences = []
for seq in sequences:
    for i in range(1, len(seq)):
        input_sequences.append(seq[:i])
        output_sequences.append(seq[i])
input_sequences = pad_sequences(input_sequences)
output_sequences = np.array(output_sequences)
# Define and Train the Model
model = Sequential([
    Embedding(input_dim=vocab_size, output_dim=50, input_length=input_sequences.shape[1]),
    LSTM(100),
    Dense(vocab_size, activation='softmax')
])
model.compile(loss='sparse_categorical_crossentropy', optimizer=Adam(learning_rate=0.01))
model.fit(input_sequences, output_sequences, epochs=50, batch_size=128)
# Generate Titles
def generate_title(seed_text, max_length=50):
    for _ in range(max_length):
        NewTitle = ""
        encoded_text = tokenizer.texts_to_sequences([seed_text])[0]
        encoded_text = pad_sequences([encoded_text], maxlen=input_sequences.shape[1])
        predicted_word_index = np.argmax(model.predict(encoded_text), axis=-1)
        predicted_word = None
        for word, index in tokenizer.word_index.items():
            if index == predicted_word_index:
                predicted_word = word
                break
        if predicted_word:
            NewTitle += " " + predicted_word
            seed_text += " " + predicted_word
        else:
            break
    return NewTitle
# make seed text
def seedText():
    submissionlList = []
    for x in range(10):
        submission = reddit.subreddit("memes").random()
        submissionlList.append(submission.title)
    user_input_text = random.choice(submissionlList)
    return user_input_text
# build model
def build_model(max_words, max_text_length):
    model = Sequential()
    model.add(Embedding(input_dim=max_words, output_dim=100, input_length=max_text_length))
    model.add(LSTM(100))
    model.add(Dense(200, activation='sigmoid'))  # Linear activation for regression
    model.compile(optimizer='SGD', loss='huber_loss')
    return model
# Variables
MAX_TEXT_LENGTH = 100
MAX_WORDS = len(tokenizer.word_index) + 1
# Load and preprocess data
data = pd.read_excel(path)
label_encoder = LabelEncoder()
y_encoded = label_encoder.fit_transform(data['Title'])
X_train, X_test, y_train, y_test = train_test_split(data, y_encoded, test_size=0.5, random_state=50, shuffle=True)
X_train_text = pad_sequences(tokenizer.texts_to_sequences(X_train['Title']), maxlen=MAX_TEXT_LENGTH)
# Build models for score, comments, and awards
score_model = build_model(MAX_WORDS, MAX_TEXT_LENGTH)
comments_model = build_model(MAX_WORDS, MAX_TEXT_LENGTH)
awards_model = build_model(MAX_WORDS, MAX_TEXT_LENGTH)
# Train and predict for score, comments, and awards
score_model.fit(X_train_text, X_train['Score'].values, batch_size=64, epochs=10, validation_split=0.5)
comments_model.fit(X_train_text, X_train['Comments'].values, batch_size=64, epochs=10, validation_split=0.5)
awards_model.fit(X_train_text, X_train['Awards'].values, batch_size=64, epochs=10, validation_split=0.5)
# Predict for user input
# Predict for user input
def get_predictions(user_input, tokenizer, max_text_length, score_model, comments_model, awards_model):
    processed_input = preprocess_text(user_input)
    # Tokenize and pad the input text
    input_sequence = pad_sequences(tokenizer.texts_to_sequences([processed_input]), maxlen=max_text_length)
    # Get predictions for score, comments, and awards
    score_prediction = round(score_model.predict(input_sequence).mean())
    comments_prediction = round(comments_model.predict(input_sequence).mean())
    awards_prediction = round(awards_model.predict(input_sequence).mean())
    # Calculate average predictions
    avg_score = round(np.mean(score_prediction))
    avg_comments = round(np.mean(comments_prediction))
    avg_awards = round(np.mean(awards_prediction))
    # Calculate confidence interval (assuming normal distribution)
    z_score = 1.96
    std_score = np.std(score_model.predict(input_sequence))
    std_comments = np.std(comments_model.predict(input_sequence))
    std_awards = np.std(awards_model.predict(input_sequence))
    confidence_interval_score = (score_prediction - z_score * std_score, score_prediction + z_score * std_score)
    confidence_interval_comments = (comments_prediction - z_score * std_comments, comments_prediction + z_score * std_comments)
    confidence_interval_awards = (awards_prediction - z_score * std_awards, awards_prediction + z_score * std_awards)
    return score_prediction, confidence_interval_score, comments_prediction, confidence_interval_comments, awards_prediction, confidence_interval_awards
# correct spelling
spell = Speller(lang='en')
lemmatizer = WordNetLemmatizer()
def correctSpelling(text):
    sentences = sent_tokenize(text)
    tokenizer = RegexpTokenizer(r'\w+')
    corrected_sentences = []
    for sentence in sentences:
        words = tokenizer.tokenize(sentence)
        corrected_words = [spell(w) for w in words]
        corrected_words = [lemmatizer.lemmatize(w, pos=wordnet.VERB) if pos_tag([w])[0][1].startswith('V') else w for w in corrected_words]
        corrected_sentence = ' '.join(corrected_words)
        corrected_sentences.append(corrected_sentence)
    preprocessed_text = ' '.join(corrected_sentences)
    return preprocessed_text
# remove unsupported characters
def remove_unsupported_characters(text):
    unsupported_char_pattern = re.compile(r'[^\x00-\x7F]+')
    cleaned_text = re.sub(unsupported_char_pattern, '', text)
    encoded_text = cleaned_text.encode('utf-8', errors='ignore')
    return encoded_text
# ------ use images to perdict Score, Comments, Awards, flair
labels_column = data['Flair'].astype(str)
custom_labels = [label for label in labels_column if remove_unsupported_characters(label) is not None]
def predict_image_from_url(image_url, model,custom_labels):
    response = requests.get(image_url, stream=True)
    img = Image.open(BytesIO(response.content))
    img = img.convert('RGB')
    img = img.resize((224, 224))
    x = image.img_to_array(img)
    x = np.expand_dims(x, axis=0)
    x = preprocess_input(x)
    predictions = model.predict(x)
    decoded_predictions = decode_predictions(predictions)
    top_prediction = decoded_predictions[0][0]
    predicted_label = top_prediction[1]
    confidence_score = top_prediction[2]
    if predicted_label in custom_labels:
        return predicted_label, confidence_score
    else:
        return "None", confidence_score
# generate random string
def generate_random_string(length=12):
    characters = string.ascii_letters
    randomString = ''.join(secrets.choice(characters) for _ in range(length))
    return randomString + ".png"
# conver to png
def download_image(url):
    output_path = generate_random_string()
    try:
        response = requests.get(url, stream=True)
        response.raise_for_status()
        image = Image.open(BytesIO(response.content))
        if image.format != "PNG":
            image = image.convert("RGBA")
            png_output = BytesIO()
            image.save(png_output, format="PNG")
            png_content = png_output.getvalue()
            with open(output_path, "wb") as png_file:
                png_file.write(png_content)
        else:
            output_path = image
        return output_path
    except Exception as e:
        print(f"Error: {e}")
        return None
# print
seed_text = seedText()
generated_title = correctSpelling(generate_title(seed_text))
image_url = "https://source.unsplash.com/random.com"
# create pos
useImage = download_image(image_url)
data_to_store = "{}\n{}".format(useImage,generated_title)
with open('data.txt', 'w') as file:
    file.write(data_to_store)
print("Data has been stored.")
